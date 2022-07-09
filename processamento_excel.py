from re import A
from sqlite3 import Row
from tkinter.tix import COLUMN
from openpyxl import Workbook
from openpyxl import load_workbook
import openpyxl
 
#Escreve ficheiro final
def escreve_excel_final(lista_colocados):
      
        from openpyxl import load_workbook
        wb = openpyxl.Workbook()
        wb.remove(wb['Sheet'])
        i=0;
      
        for key,value in lista_colocados.items(): 
                folha = wb.create_sheet(key)
                if key != "fora":
                        write_cabecalho(folha)
                        i=0;
                        for x in value:
                                a=1;
                                for item in x:
                                        folha.cell(i+2,a,str(item))
                                        a = a + 1
                                i = i + 1
                else:
                        i = 1
                        folha.cell(i,1,"Nº Processo Matricula")
                        for x in value:
                                i = i + 1
                                folha.cell(i,1,str(x))
                                
        wb.save("Distribuicao_Final.xlsx")
    
#Abrir ficheiro codigos_postais()
def open_excel_reading_codPostal():
        from openpyxl import load_workbook
        wb = load_workbook(filename = './CODIGOS_POSTAIS.xlsx')
        return wb['Folha1']

#Abrir ficheiro ipss()
def open_excel_reading_ipss():
        from openpyxl import load_workbook
        wb = load_workbook(filename = './IPSS.xlsx')
        return wb['Folha1']

# Abrir ficheiro excel do portal de matrículas
def open_excel_reading_matriculas():
        from openpyxl import load_workbook
        wb = load_workbook(filename = './exportacao_portal_seriacao.xlsx')
        return wb['Seriação']

# Abrir ficheiro excel das escolas
def open_excel_reading_escolas():
        from openpyxl import load_workbook
        wb = load_workbook(filename = './escolas.xlsx')
        return wb['escolas']

def write_cabecalho(matriculas):
      
        matriculas.cell(1,1,"Nºprocesso")          
        matriculas.cell(1,2,"Nome do aluno")
        matriculas.cell(1,3,"Data Nascimento")
        matriculas.cell(1,4,"Cod Postal") 
        matriculas.cell(1,5,"Pais Menores") 
        matriculas.cell(1,6,"NEE")
        matriculas.cell(1,7,"Freq Creche")
        matriculas.cell(1,8,"Cod Escola Anterior")
        matriculas.cell(1,9,"Cod Escola 1")
        matriculas.cell(1,10,"Nome Escola 1")
        matriculas.cell(1,11,"Irmãos Escola 1")
        matriculas.cell(1,12,"EE Emprego Escola 1")
        matriculas.cell(1,13,"Cod Escola 2")
        matriculas.cell(1,14,"Nome Escola 2")
        matriculas.cell(1,15,"Irmãos Escola 2")
        matriculas.cell(1,16,"EE Emprego Escola 2")
        matriculas.cell(1,17,"Cod Escola 3")
        matriculas.cell(1,18,"Nome Escola 3")
        matriculas.cell(1,19,"Irmãos Escola 3")
        matriculas.cell(1,20,"EE Emprego Escola 3")
        matriculas.cell(1,21,"Cod Escola 4")
        matriculas.cell(1,22,"Nome Escola 4")
        matriculas.cell(1,23,"Irmãos Escola 4")
        matriculas.cell(1,24,"EE Emprego Escola 4")
        matriculas.cell(1,25,"Cod Escola 5")
        matriculas.cell(1,26,"Nome Escola 5")
        matriculas.cell(1,27,"Irmãos Escola 5")
        matriculas.cell(1,28,"EE Emprego Escola 5") 
        matriculas.cell(1,29,"Escalão") 
        matriculas.cell(1,30,"Reduz Turma")  


def write_result(final_schedule,nome_folha):
    from openpyxl import load_workbook
    wb = openpyxl.Workbook()
    matriculas = wb.create_sheet('MATRICULAS')
    wb.remove(wb['Sheet'])
    i=0;
    
    write_cabecalho(matriculas)   
    
    for key,value in final_schedule.items(): 
        
        matriculas.cell(i+2,1,str(key))       
        a=2;
        for x in value:
                matriculas.cell(i+2,a,str(x))
                a = a + 1
        i = i + 1
    wb.save(nome_folha)
    
    
